from openpyxl import load_workbook, Workbook
import tkinter as tk


def merge_click(xl_list):
    master_book = Workbook()
    master_sheet = master_book.active
    master_sheet.title = 'Master Sheet'
    master_row_counter = master_sheet.max_row + 1

    for sheet in xl_list:
        current_file = load_workbook(sheet)
        current_sheet = current_file.active
        for row in range(1, current_sheet.max_row + 1):
            print('\n')
            print('Row ', row, "data :")

            # sets up the for-loop to get data from the source spreadsheet column
            for col in range(1, current_sheet.max_column + 1):
                # setting variable to store the value of the current cell iteration
                cell_obj = current_sheet.cell(row=row, column=col)
                print(cell_obj, end=' ')
                print(cell_obj.value, end=' \n')
                # setting up the cell at the end of the master workbook to write data from the source
                master_cell = master_sheet.cell(row=master_row_counter, column=col)
                # assigning the value of the source cell to the master cell at the current position of the master cell
                master_cell.value = cell_obj.value

            # incrimenting the master row counter in the first loop to advance to the next row upon next iteration
            master_row_counter += 1
            print(f"Row Counter Is At: {master_row_counter}")

    def get_master_name():

        def submit_merge(entry_str):
            master_name = entry_str
            master_book.save(master_name)
            pop_up_window.destroy()
            naming_window.destroy()

        naming_window = tk.Tk()

        naming_label = tk.Label(naming_window, text='Enter the name of the master file (include .xlsx)')
        naming_label.pack()

        pop_up_name_entry = tk.Entry(naming_window)
        pop_up_name_entry.pack()

        confirm_merge_button = tk.Button(naming_window, text="Confirm files to merge", command=lambda: submit_merge(pop_up_name_entry.get()))
        confirm_merge_button.pack()

        naming_window.mainloop()

    pop_up_window = tk.Tk()
    pop_up_window.geometry('400x300')

    pop_up_greeting = tk.Label(pop_up_window, text="Merge these files?")
    pop_up_greeting.pack()

    pop_up_output = tk.Text(pop_up_window, height=8)
    pop_up_output.pack()

    for count, file in enumerate(xl_list):
        current_filename = file.split("/")[-1] + '\n'
        # print(file.split("/")[-1])
        pop_up_output.insert(f"{count + 1}.0", current_filename)

    pop_up_button_close = tk.Button(pop_up_window, text="Close", command=pop_up_window.destroy)
    pop_up_button_close.pack()

    pop_up_button_merge = tk.Button(pop_up_window, text="Merge", command=get_master_name)
    pop_up_button_merge.pack()

    pop_up_window.mainloop()
# xl_dir = input("Enter a directory to merge all xl files inside of: ")
# xl_list = get_all_xl_files_in_dir(xl_dir)
# merge_click(xl_list)
# /home/scrant/PycharmProjects/Scrap
